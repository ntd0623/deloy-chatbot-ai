from flask import Flask, request, jsonify
from flask_cors import CORS
import json
import os
import openpyxl
from sentence_transformers import SentenceTransformer
from sklearn.metrics.pairwise import cosine_similarity
import numpy as np

app = Flask(__name__)
CORS(app)  # Cho phép gọi API từ frontend

DATA_FILE = "data.json"
EXCEL_FILE = "data.xlsx"
UPLOAD_FOLDER = "uploads"  # Thư mục để lưu file upload
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Load model
model = SentenceTransformer("all-MiniLM-L6-v2")

# Load data từ file JSON
if os.path.exists(DATA_FILE):
    with open(DATA_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)
else:
    data = {}

# --- Xử lý dữ liệu ---
def get_embeddings():
    if not data:
        return [], []
    questions = list(data.keys())
    vectors = model.encode(questions)
    return questions, vectors

questions, vectors = get_embeddings()

def save_data():
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=4)
    global questions, vectors
    questions, vectors = get_embeddings()

def find_best_match(user_input, threshold=0.6):
    if not data:
        return None
    user_vec = model.encode([user_input])
    sims = cosine_similarity(user_vec, vectors)[0]
    best_idx = np.argmax(sims)
    if sims[best_idx] >= threshold:
        return questions[best_idx]
    return None

def train_from_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    added = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            q = str(row[0]).lower().strip()
            a = str(row[1]).strip()
            if q not in data:
                data[q] = a
                added += 1
    save_data()
    return added

# --- API Routes ---
@app.route("/chat", methods=["POST"])
def chat():
    user_input = request.json.get("message", "").lower().strip()
    match = find_best_match(user_input)
    if match:
        return jsonify({"reply": data[match]})
    else:
        return jsonify({"reply": "Mình chưa biết câu trả lời. Hãy upload file Excel để dạy mình nhé!"})

@app.route("/upload", methods=["POST"])
def upload_excel():
    if "file" not in request.files:
        return jsonify({"error": "❌ Không có file nào được gửi!"}), 400

    file = request.files["file"]

    if not file.filename.endswith(".xlsx"):
        return jsonify({"error": "❌ File không hợp lệ, chỉ chấp nhận .xlsx"}), 400

    try:
        # Lưu file Excel vào thư mục uploads
        file_path = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(file_path)

        # Đọc Excel và convert sang JSON
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        excel_data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                q = str(row[0]).strip()
                a = str(row[1]).strip()
                excel_data.append({"question": q, "answer": a})

        # Lưu JSON vào thư mục uploads (cùng tên Excel nhưng .json)
        json_filename = file.filename.replace(".xlsx", ".json")
        json_path = os.path.join(UPLOAD_FOLDER, json_filename)
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(excel_data, f, ensure_ascii=False, indent=4)

        # Train luôn với dữ liệu Excel
        added = train_from_excel(file_path)

        return jsonify({
            "message": f"✅ Upload & train thành công: {added} câu mới từ {file.filename}",
            "json_file": json_filename
        })
    except Exception as e:
        return jsonify({"error": f"❌ Lỗi xử lý file: {str(e)}"}), 500

if __name__ == "__main__":
    app.run(debug=True)
